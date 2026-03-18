#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
飞书多维表格上传模块
将爬取结果上传到飞书多维表格
"""

import os
import json
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime

import requests
from openpyxl import load_workbook

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class FeishuUploader:
    """飞书上传器"""

    # 飞书 API 端点
    BASE_URL = "https://open.feishu.cn/open-apis"

    def __init__(self, config_path: Optional[str] = None):
        """
        初始化飞书上传器

        Args:
            config_path: 配置文件路径，默认为 ~/.openclaw/feishu/feishu_config.json
        """
        self.config_path = config_path or os.path.expanduser("~/.openclaw/feishu/feishu_config.json")
        self.config = self._load_config()
        self.access_token = None

    def _load_config(self) -> Dict[str, Any]:
        """加载配置"""
        default_config = {
            "app_id": "",
            "app_secret": "",
            "bitable_id": "",
            "table_id": "",
            "view_id": ""
        }

        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    # 合并默认值
                    for key, value in default_config.items():
                        if key not in config:
                            config[key] = value
                    return config
            except Exception as e:
                logger.warning(f"加载配置失败: {e}，使用默认配置")

        return default_config

    def _save_config(self):
        """保存配置"""
        os.makedirs(os.path.dirname(self.config_path), exist_ok=True)
        with open(self.config_path, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, indent=2, ensure_ascii=False)

    def get_access_token(self) -> str:
        """获取访问令牌"""
        if self.access_token:
            return self.access_token

        app_id = self.config.get("app_id")
        app_secret = self.config.get("app_secret")

        if not app_id or not app_secret:
            raise ValueError("未配置 app_id 或 app_secret，请先配置飞书凭据")

        url = f"{self.BASE_URL}/auth/v3/tenant_access_token/internal"
        payload = {
            "app_id": app_id,
            "app_secret": app_secret
        }

        response = requests.post(url, json=payload)
        result = response.json()

        if result.get("code") != 0:
            raise Exception(f"获取访问令牌失败: {result.get('msg')}")

        self.access_token = result.get("tenant_access_token")
        return self.access_token

    def upload_file(self, file_path: str, parent_type: str = "explorer") -> Dict[str, Any]:
        """
        上传文件到飞书

        Args:
            file_path: 文件路径
            parent_type: 父节点类型 (explorer/wiki)

        Returns:
            上传结果，包含 file_token 和 url
        """
        token = self.get_access_token()
        file_name = os.path.basename(file_path)

        # 准备文件
        files = {
            "file": (file_name, open(file_path, "rb").read())
        }
        data = {
            "parent_type": parent_type,
            "parent_node": self.config.get("parent_node", "")
        }

        headers = {
            "Authorization": f"Bearer {token}"
        }

        url = f"{self.BASE_URL}/drive/v1/files/upload_all"
        response = requests.post(url, headers=headers, files=files, data=data)
        result = response.json()

        if result.get("code") != 0:
            raise Exception(f"文件上传失败: {result.get('msg')}")

        file_token = result.get("data", {}).get("file_token")

        # 获取文件链接
        return self.get_file_url(file_token)

    def get_file_url(self, file_token: str) -> Dict[str, Any]:
        """获取文件访问链接"""
        token = self.get_access_token()

        url = f"{self.BASE_URL}/drive/v1/files/{file_token}/preview"
        headers = {
            "Authorization": f"Bearer {token}"
        }

        response = requests.get(url, headers=headers)
        result = response.json()

        if result.get("code") != 0:
            raise Exception(f"获取文件链接失败: {result.get('msg')}")

        return {
            "file_token": file_token,
            "url": result.get("data", {}).get("url", "")
        }

    def read_excel_data(self, excel_path: str) -> List[Dict[str, Any]]:
        """
        读取 Excel 数据

        Args:
            excel_path: Excel 文件路径

        Returns:
            数据行列表
        """
        wb = load_workbook(excel_path)
        ws = wb.active

        # 获取表头
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # 读取数据行
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            rows.append(row_dict)

        return rows

    def create_bitable_record(self, table_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        创建多维表格记录

        Args:
            table_data: 表格数据

        Returns:
            创建结果
        """
        token = self.get_access_token()
        app_token = self.config.get("app_token")
        table_id = self.config.get("table_id")

        if not app_token or not table_id:
            raise ValueError("未配置 app_token 或 table_id")

        url = f"{self.BASE_URL}/bitable/v1/apps/{app_token}/tables/{table_id}/records"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }

        response = requests.post(url, headers=headers, json=table_data)
        result = response.json()

        if result.get("code") != 0:
            raise Exception(f"创建记录失败: {result.get('msg')}")

        return result.get("data", {})

    def upload_crawl_result(self, crawl_result: Dict[str, Any]) -> Dict[str, Any]:
        """
        上传爬取结果到飞书

        Args:
            crawl_result: 爬取结果字典
                {
                    "keyword": str,
                    "excel_file": str,
                    "images_dir": str,
                    "video_files": list,
                    "count": int,
                    "timestamp": str
                }

        Returns:
            上传结果
        """
        logger.info(f"开始上传爬取结果: {crawl_result.get('keyword')}")

        result = {
            "success": False,
            "records_created": 0,
            "files_uploaded": [],
            "url": "",
            "error": None
        }

        try:
            # 1. 上传 Excel 文件
            excel_file = crawl_result.get("excel_file")
            if excel_file and os.path.exists(excel_file):
                excel_result = self.upload_file(excel_file)
                result["files_uploaded"].append({
                    "type": "excel",
                    "name": os.path.basename(excel_file),
                    "url": excel_result.get("url")
                })
                logger.info(f"Excel 上传成功: {excel_result.get('url')}")

            # 2. 上传图片（如果有）
            images_dir = crawl_result.get("images_dir")
            if images_dir and os.path.exists(images_dir):
                for img_file in Path(images_dir).rglob("*.jpg"):
                    try:
                        img_result = self.upload_file(str(img_file))
                        result["files_uploaded"].append({
                            "type": "image",
                            "name": img_file.name,
                            "url": img_result.get("url")
                        })
                    except Exception as e:
                        logger.warning(f"图片上传失败 {img_file}: {e}")

            # 3. 上传视频（如果有）
            for video_file in crawl_result.get("video_files", []):
                if os.path.exists(video_file):
                    try:
                        video_result = self.upload_file(video_file)
                        result["files_uploaded"].append({
                            "type": "video",
                            "name": os.path.basename(video_file),
                            "url": video_result.get("url")
                        })
                    except Exception as e:
                        logger.warning(f"视频上传失败 {video_file}: {e}")

            # 4. 创建多维表格记录（如果配置了）
            try:
                if excel_file and os.path.exists(excel_file):
                    rows = self.read_excel_data(excel_file)

                    for row in rows[:10]:  # 限制前10条作为示例
                        # 转换为飞书字段格式
                        record = self._convert_to_feishu_record(row, crawl_result)
                        self.create_bitable_record(record)
                        result["records_created"] += 1

            except Exception as e:
                logger.warning(f"创建多维表格记录失败: {e}")

            result["success"] = True
            result["url"] = result["files_uploaded"][0].get("url") if result["files_uploaded"] else ""

            logger.info(f"上传完成: {result['records_created']} 条记录, {len(result['files_uploaded'])} 个文件")

        except Exception as e:
            logger.error(f"上传失败: {e}")
            result["error"] = str(e)

        return result

    def _convert_to_feishu_record(self, row_data: Dict[str, Any], crawl_result: Dict[str, Any]) -> Dict[str, Any]:
        """
        将爬取数据转换为飞书记录格式

        Args:
            row_data: 行数据
            crawl_result: 爬取结果

        Returns:
            飞书记录格式
        """
        # 基本字段映射（与飞书多维表格字段完全对应）
        fields = {}

        # 文本字段映射
        field_mapping = {
            "note_id": "笔记ID",
            "title": "标题",
            "author": "作者",
            "content": "正文",
            "tags": "标签",
            "publish_time": "发布时间",
            "ip_region": "IP地区",
            "note_type": "笔记类型",
            "note_link": "笔记的链接"
        }

        for key, label in field_mapping.items():
            if key in row_data and row_data[key] is not None:
                fields[label] = row_data[key]

        # 数值字段
        if "like_count" in row_data and row_data["like_count"] is not None:
            fields["点赞数"] = int(row_data["like_count"])
        if "collect_count" in row_data and row_data["collect_count"] is not None:
            fields["收藏数"] = int(row_data["collect_count"])
        if "comment_count" in row_data and row_data["comment_count"] is not None:
            fields["评论数"] = int(row_data["comment_count"])

        # 添加元数据
        fields["来源关键词"] = crawl_result.get("keyword", "")
        # 飞书时间字段需要 Unix 时间戳（毫秒）
        fields["爬取时间"] = int(datetime.now().timestamp() * 1000)

        return {"fields": fields}

    def configure(self, app_id: str = None, app_secret: str = None,
                  bitable_id: str = None, table_id: str = None):
        """
        配置飞书凭据

        Args:
            app_id: 飞书应用 ID
            app_secret: 飞书应用密钥
            bitable_id: 多维表格 ID
            table_id: 表格 ID
        """
        if app_id:
            self.config["app_id"] = app_id
        if app_secret:
            self.config["app_secret"] = app_secret
        if bitable_id:
            self.config["bitable_id"] = bitable_id
        if table_id:
            self.config["table_id"] = table_id

        self._save_config()
        logger.info("配置已保存")


# ============== 命令行接口 ==============

def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(description="飞书上传工具")
    parser.add_argument("excel_file", help="Excel 文件路径")
    parser.add_argument("--configure", action="store_true", help="配置模式")

    args = parser.parse_args()

    uploader = FeishuUploader()

    if args.configure:
        # 配置模式
        print("飞书上传配置")
        app_id = input("App ID: ").strip()
        app_secret = input("App Secret: ").strip()
        bitable_id = input("多维表格 ID (可选): ").strip()
        table_id = input("表格 ID (可选): ").strip()

        uploader.configure(
            app_id=app_id or None,
            app_secret=app_secret or None,
            bitable_id=bitable_id or None,
            table_id=table_id or None
        )
        print("配置已保存")
        return

    # 上传模式
    if not os.path.exists(args.excel_file):
        print(f"文件不存在: {args.excel_file}")
        return

    result = uploader.upload_file(args.excel_file)
    print(f"上传成功: {result.get('url')}")


if __name__ == "__main__":
    main()
